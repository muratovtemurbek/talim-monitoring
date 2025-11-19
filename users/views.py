from rest_framework import status, generics
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate, get_user_model
from django.db import models
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncMonth, TruncDay
from datetime import datetime, timedelta
from .serializers import (
    UserSerializer,
    RegisterSerializer,
    LoginSerializer,
    ChangePasswordSerializer,
)

User = get_user_model()


class RegisterView(generics.CreateAPIView):
    """Ro'yxatdan o'tish"""
    queryset = User.objects.all()
    permission_classes = [AllowAny]
    serializer_class = RegisterSerializer


class LoginView(APIView):
    """Login"""
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        username = serializer.validated_data['username']
        password = serializer.validated_data['password']

        user = authenticate(username=username, password=password)

        if user is None:
            return Response(
                {'error': 'Login yoki parol noto\'g\'ri'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        if not user.is_active:
            return Response(
                {'error': 'Akkaunt faol emas'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        refresh = RefreshToken.for_user(user)

        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'user': UserSerializer(user).data
        })


class LogoutView(APIView):
    """Logout"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh")
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response({'message': 'Muvaffaqiyatli chiqildi'})
        except Exception:
            return Response(
                {'error': 'Token bekor qilishda xatolik'},
                status=status.HTTP_400_BAD_REQUEST
            )


class UserProfileView(generics.RetrieveUpdateAPIView):
    """Profil"""
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    def get_object(self):
        return self.request.user


class ChangePasswordView(APIView):
    """Parol o'zgartirish"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        old_password = serializer.validated_data['old_password']
        new_password = serializer.validated_data['new_password']

        if not user.check_password(old_password):
            return Response(
                {'error': 'Joriy parol noto\'g\'ri'},
                status=status.HTTP_400_BAD_REQUEST
            )

        user.set_password(new_password)
        user.save()

        return Response({'message': 'Parol muvaffaqiyatli o\'zgartirildi'})


class UserListView(generics.ListCreateAPIView):
    """Foydalanuvchilar ro'yxati (Admin)"""
    permission_classes = [IsAdminUser]
    serializer_class = UserSerializer
    queryset = User.objects.all()


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Foydalanuvchi detallar (Admin)"""
    permission_classes = [IsAdminUser]
    serializer_class = UserSerializer
    queryset = User.objects.all()


# ANALYTICS VIEWS

class DashboardStatsView(APIView):
    """Dashboard statistikasi"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from teachers.models import Teacher, TeacherActivity
        from materials.models import Material
        from videos.models import Video
        from schools.models import School
        from consultations.models import Consultation

        user = request.user

        if user.is_superuser or user.role == 'superadmin':
            stats = {
                'total_users': User.objects.count(),
                'total_teachers': Teacher.objects.count(),
                'total_schools': School.objects.count(),
                'total_materials': Material.objects.count(),
                'total_videos': Video.objects.count(),
                'pending_materials': Material.objects.filter(is_approved=False).count(),
                'pending_videos': Video.objects.filter(is_approved=False).count(),
                'total_consultations': Consultation.objects.count(),
            }
        elif user.role == 'admin':
            school = School.objects.filter(director=user).first()
            if school:
                stats = {
                    'school_name': school.name,
                    'total_teachers': Teacher.objects.filter(school=school).count(),
                    'total_materials': Material.objects.filter(teacher__school=school).count(),
                    'total_videos': Video.objects.filter(teacher__school=school).count(),
                    'pending_materials': Material.objects.filter(teacher__school=school, is_approved=False).count(),
                    'pending_videos': Video.objects.filter(teacher__school=school, is_approved=False).count(),
                }
            else:
                stats = {}
        elif user.role == 'teacher':
            try:
                teacher = Teacher.objects.get(user=user)
                stats = {
                    'total_points': teacher.total_points,
                    'monthly_points': teacher.monthly_points,
                    'level': teacher.level,
                    'total_materials': Material.objects.filter(teacher=teacher).count(),
                    'approved_materials': Material.objects.filter(teacher=teacher, is_approved=True).count(),
                    'total_videos': Video.objects.filter(teacher=teacher).count(),
                    'approved_videos': Video.objects.filter(teacher=teacher, is_approved=True).count(),
                    'total_activities': TeacherActivity.objects.filter(teacher=teacher).count(),
                }
            except Teacher.DoesNotExist:
                stats = {}
        else:
            stats = {}

        return Response(stats)


class MonthlyActivityView(APIView):
    """Oylik faoliyat statistikasi"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from teachers.models import Teacher, TeacherActivity
        from django.db.models import Count, Sum
        from django.db.models.functions import TruncMonth

        six_months_ago = datetime.now() - timedelta(days=180)

        if request.user.role == 'teacher':
            try:
                teacher = Teacher.objects.get(user=request.user)
                activities = TeacherActivity.objects.filter(
                    teacher=teacher,
                    date__gte=six_months_ago
                ).annotate(
                    month=TruncMonth('date')
                ).values('month').annotate(
                    count=Count('id'),
                    total_points=Sum('points')
                ).order_by('month')

                return Response(list(activities))
            except Teacher.DoesNotExist:
                return Response([])

        return Response([])


class TopPerformersView(APIView):
    """Top o'qituvchilar"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from teachers.models import Teacher

        limit = request.query_params.get('limit', 10)

        if request.user.role == 'admin':
            teachers = Teacher.objects.filter(
                school__director=request.user
            ).order_by('-total_points')[:int(limit)]
        else:
            teachers = Teacher.objects.order_by('-total_points')[:int(limit)]

        from teachers.serializers import TeacherSerializer
        serializer = TeacherSerializer(teachers, many=True)
        return Response(serializer.data)


class BulkApprovalView(APIView):
    """Ko'plab materiallarni tasdiqlash"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role not in ['admin', 'superadmin']:
            return Response(
                {'error': 'Faqat admin tasdiqlashi mumkin'},
                status=status.HTTP_403_FORBIDDEN
            )

        from materials.models import Material
        from videos.models import Video

        material_ids = request.data.get('material_ids', [])
        video_ids = request.data.get('video_ids', [])
        action = request.data.get('action', 'approve')

        if action == 'approve':
            materials = Material.objects.filter(id__in=material_ids)
            for material in materials:
                material.is_approved = True
                material.save()
                material.teacher.total_points += 10
                material.teacher.monthly_points += 10
                material.teacher.update_level()

            videos = Video.objects.filter(id__in=video_ids)
            for video in videos:
                video.is_approved = True
                video.save()
                video.teacher.total_points += 15
                video.teacher.monthly_points += 15
                video.teacher.update_level()

            message = f"{len(materials)} material va {len(videos)} video tasdiqlandi"

        elif action == 'reject':
            Material.objects.filter(id__in=material_ids).delete()
            Video.objects.filter(id__in=video_ids).delete()
            message = f"{len(material_ids)} material va {len(video_ids)} video rad etildi"

        else:
            return Response(
                {'error': 'Noto\'g\'ri action'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({'message': message})


class PendingApprovalsView(APIView):
    """Kutilayotgan tasdiqlar"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role not in ['admin', 'superadmin']:
            return Response(
                {'error': 'Faqat admin ko\'rishi mumkin'},
                status=status.HTTP_403_FORBIDDEN
            )

        from materials.models import Material
        from videos.models import Video
        from materials.serializers import MaterialSerializer
        from videos.serializers import VideoSerializer

        if request.user.role == 'admin':
            materials = Material.objects.filter(
                teacher__school__director=request.user,
                is_approved=False
            )
            videos = Video.objects.filter(
                teacher__school__director=request.user,
                is_approved=False
            )
        else:
            materials = Material.objects.filter(is_approved=False)
            videos = Video.objects.filter(is_approved=False)

        return Response({
            'materials': MaterialSerializer(materials, many=True).data,
            'videos': VideoSerializer(videos, many=True).data,
        })


# EXPORT VIEWS

class ExportRatingsExcelView(APIView):
    """Reytingni Excel ga export qilish"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from ratings.models import TeacherRating

        wb = Workbook()
        ws = wb.active
        ws.title = "O'qituvchilar Reytingi"

        ws['A1'] = "O'qituvchilar Reytingi"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')

        headers = ['#', 'O\'qituvchi', 'Maktab', 'Ball', 'O\'rin']
        ws.append([])
        ws.append(headers)

        for cell in ws[3]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        ratings = TeacherRating.objects.select_related('teacher').order_by('rank')[:50]
        for idx, rating in enumerate(ratings, 1):
            ws.append([
                idx,
                rating.teacher.user.get_full_name(),
                rating.teacher.school.name,
                rating.total_points,
                rating.rank
            ])

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reyting_{datetime.now().strftime("%Y%m%d")}.xlsx'

        wb.save(response)
        return response


class ExportRatingsPDFView(APIView):
    """Reytingni PDF ga export qilish"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from ratings.models import TeacherRating

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()

        title = Paragraph("O'qituvchilar Reytingi", styles['Title'])
        elements.append(title)
        elements.append(Paragraph("<br/><br/>", styles['Normal']))

        data = [['#', 'O\'qituvchi', 'Maktab', 'Ball', 'O\'rin']]

        ratings = TeacherRating.objects.select_related('teacher').order_by('rank')[:50]
        for idx, rating in enumerate(ratings, 1):
            data.append([
                str(idx),
                rating.teacher.user.get_full_name(),
                rating.teacher.school.name[:30],
                str(rating.total_points),
                str(rating.rank)
            ])

        table = Table(data, colWidths=[0.5 * inch, 2 * inch, 3 * inch, 1 * inch, 1 * inch])

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=reyting_{datetime.now().strftime("%Y%m%d")}.pdf'

        return response


class ExportLessonAnalysisExcelView(APIView):
    """Dars tahlilini Excel ga export"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from lesson_analysis.models import LessonAnalysis

        wb = Workbook()
        ws = wb.active
        ws.title = "Dars Tahlillari"

        ws['A1'] = "Dars Tahlillari Hisoboti"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')

        headers = ['#', 'O\'qituvchi', 'Tahlilchi', 'Mavzu', 'Fan', 'Sinf', 'Baho', 'Sana']
        ws.append([])
        ws.append(headers)

        for cell in ws[3]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        analyses = LessonAnalysis.objects.select_related('teacher', 'analyzer').filter(
            status='approved'
        ).order_by('-lesson_date')[:100]

        for idx, analysis in enumerate(analyses, 1):
            ws.append([
                idx,
                analysis.teacher.user.get_full_name(),
                analysis.analyzer.user.get_full_name(),
                analysis.topic,
                analysis.subject,
                analysis.grade,
                float(analysis.overall_rating),
                analysis.lesson_date.strftime('%Y-%m-%d')
            ])

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=dars_tahlili_{datetime.now().strftime("%Y%m%d")}.xlsx'

        wb.save(response)
        return response


# CHART DATA VIEWS

class ChartDataView(APIView):
    """Grafiklar uchun ma'lumotlar"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        chart_type = request.query_params.get('type', 'monthly_activity')

        if chart_type == 'monthly_activity':
            return self.get_monthly_activity(request)
        elif chart_type == 'subject_distribution':
            return self.get_subject_distribution(request)
        elif chart_type == 'teacher_comparison':
            return self.get_teacher_comparison(request)
        elif chart_type == 'school_comparison':
            return self.get_school_comparison(request)
        elif chart_type == 'growth_trend':
            return self.get_growth_trend(request)
        else:
            return Response({'error': 'Invalid chart type'}, status=400)

    def get_monthly_activity(self, request):
        """Oylik faoliyat grafigi"""
        from teachers.models import TeacherActivity

        six_months_ago = datetime.now() - timedelta(days=180)

        if request.user.role == 'teacher':
            from teachers.models import Teacher
            try:
                teacher = Teacher.objects.get(user=request.user)
                activities = TeacherActivity.objects.filter(
                    teacher=teacher,
                    date__gte=six_months_ago
                ).annotate(
                    month=TruncMonth('date')
                ).values('month').annotate(
                    materials=Count('id', filter=Q(activity_type='material_upload')),
                    videos=Count('id', filter=Q(activity_type='video_upload')),
                    analyses=Count('id', filter=Q(activity_type='lesson_analysis')),
                    total_points=Sum('points')
                ).order_by('month')

                return Response(list(activities))
            except Teacher.DoesNotExist:
                return Response([])

        # Admin uchun umumiy statistika
        activities = TeacherActivity.objects.filter(
            date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            materials=Count('id', filter=Q(activity_type='material_upload')),
            videos=Count('id', filter=Q(activity_type='video_upload')),
            analyses=Count('id', filter=Q(activity_type='lesson_analysis')),
            total_points=Sum('points')
        ).order_by('month')

        return Response(list(activities))

    def get_subject_distribution(self, request):
        """Fanlar bo'yicha taqsimot"""
        from materials.models import Material

        distribution = Material.objects.values('subject').annotate(
            count=Count('id')
        ).order_by('-count')

        return Response(list(distribution))

    def get_teacher_comparison(self, request):
        """O'qituvchilar taqqoslash (Top 10)"""
        from teachers.models import Teacher

        teachers = Teacher.objects.order_by('-total_points')[:10]

        data = [{
            'name': t.user.get_full_name(),
            'total_points': t.total_points,
            'monthly_points': t.monthly_points,
            'materials_count': t.material_set.count(),
            'videos_count': t.video_set.count(),
        } for t in teachers]

        return Response(data)

    def get_school_comparison(self, request):
        """Maktablar taqqoslash"""
        from schools.models import School

        schools = School.objects.annotate(
            teacher_count=Count('teacher'),
            total_points=Sum('teacher__total_points')
        ).order_by('-total_points')[:10]

        data = [{
            'name': s.name,
            'teacher_count': s.teacher_count,
            'total_points': s.total_points or 0,
        } for s in schools]

        return Response(data)

    def get_growth_trend(self, request):
        """O'sish trendi (oxirgi 30 kun)"""
        from teachers.models import TeacherActivity

        thirty_days_ago = datetime.now() - timedelta(days=30)

        daily_stats = TeacherActivity.objects.filter(
            date__gte=thirty_days_ago
        ).annotate(
            day=TruncDay('date')
        ).values('day').annotate(
            total_activities=Count('id'),
            total_points=Sum('points')
        ).order_by('day')

        return Response(list(daily_stats))


class AnalyticsOverviewView(APIView):
    """Umumiy tahliliy ko'rsatkichlar"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from teachers.models import Teacher, TeacherActivity
        from materials.models import Material
        from videos.models import Video
        from lesson_analysis.models import LessonAnalysis
        from schools.models import School

        thirty_days_ago = datetime.now() - timedelta(days=30)

        overview = {
            'total_teachers': Teacher.objects.count(),
            'active_teachers': Teacher.objects.filter(
                teacheractivity__date__gte=thirty_days_ago
            ).distinct().count(),
            'total_schools': School.objects.count(),
            'total_materials': Material.objects.count(),
            'approved_materials': Material.objects.filter(is_approved=True).count(),
            'total_videos': Video.objects.count(),
            'approved_videos': Video.objects.filter(is_approved=True).count(),
            'total_analyses': LessonAnalysis.objects.count(),
            'approved_analyses': LessonAnalysis.objects.filter(status='approved').count(),
            'total_points_distributed': Teacher.objects.aggregate(
                total=Sum('total_points')
            )['total'] or 0,
            'monthly_activities': TeacherActivity.objects.filter(
                date__gte=thirty_days_ago
            ).count(),
        }

        return Response(overview)